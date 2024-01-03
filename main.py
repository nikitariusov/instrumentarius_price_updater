from openpyxl import load_workbook
from parameters.main import get_settings
from product.main import load_xml_file, get_price
from utils.utils import open_file
from excel_update.main import update
import datetime

price = {}
params = get_settings()
data = datetime.datetime.now()

file = open_file('Excel')
wb = load_workbook(file)
sheet = wb['Export Products Sheet']
max_row = sheet.max_row


if __name__ == "__main__":
    for param in params:
        try:
            print(f'\nStart parsing: {param.company}...')
            tree = load_xml_file(param.link)
            get_price(price, tree, param)
        except:
            print("Some error...")
            continue

    print("Обновление цен...")
    update(price, sheet, max_row)

    file_name = f'{data.day}.{data.month}.{data.day}_{data.hour}-{data.minute}_update.xlsx'
    wb.save(file_name)
    print(f"Сохранен файл: {file_name}")
    input("Для выхода нажми Enter...")
