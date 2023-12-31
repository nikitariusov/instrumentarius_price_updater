from openpyxl.styles import PatternFill


def update(price, sheet, max_row):
    id_cell = 'A'
    price_cell = 'I'
    available_cell = 'P'
    start_row = 2
    yellow_fill = PatternFill(start_color='00FFFF00',  # стиль заливки измнененных ячеек.
                              end_color='00FFFF00',
                              fill_type='solid')
    red_fill = PatternFill(start_color='00FF0000',  # стиль заливки не измененных ячеек.
                              end_color='00FF0000',
                              fill_type='solid')

    for row in range(start_row, max_row + 1):
        cell_id = sheet[f'{id_cell}{row}']
        cell_price = sheet[f'{price_cell}{row}']
        cell_available = sheet[f'{available_cell}{row}']
        product_id = cell_id.value

        set_price(cell_id, price, product_id, cell_price, yellow_fill, red_fill)
        set_available(price, product_id, cell_available, yellow_fill, red_fill)


def set_price(cell_id, price, product_id, cell_price, yellow_fill, red_fill):
    if product_id in price and cell_price.value != price[product_id].price:
        cell_price.value = price[product_id].price
        cell_price.fill = yellow_fill
    elif product_id not in price:
        cell_id.fill = red_fill
        cell_price.fill = red_fill
    else:
        pass


def set_available(price, product_id, cell_available, yellow_fill, red_fill):
    if product_id in price:
        if price[product_id].available == "true":
            available = '+'
        else:
            available = '-'
        if cell_available.value != available:
            cell_available.value = available
            cell_available.fill = yellow_fill
    elif product_id not in price:
        cell_available.fill = red_fill
        cell_available.value = '-'
